import os
import asyncio
import logging
from datetime import datetime
import random
from openpyxl import load_workbook
from aiogram import Bot, Dispatcher
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from dotenv import load_dotenv

# Загружаем секретные переменные из файла .env
load_dotenv()

# === Настройки ===
BOT_TOKEN = os.getenv("BOT_TOKEN")
HR_CHAT_ID = int(os.getenv("HR_CHAT_ID", 0))
# Путь к файлу можно задать в .env, либо использовать по умолчанию "employees.xlsx"
EXCEL_PATH = os.getenv("EXCEL_PATH", "employees.xlsx")

# === Пулы поздравлений ===
WISHES_FORMAL = [
    "Уважаемый(ая) {name}! От лица всей команды поздравляем Вас с днем рождения. Желаем успехов в управлении, масштабных свершений, надежной команды и крепкого здоровья!",
    "С днем рождения, {name}! Желаем Вам достижения всех стратегических целей, неиссякаемой энергии для новых проектов и процветания. Пусть работа приносит только удовлетворение.",
    "{name}, поздравляем с праздником! Желаем стабильного роста Вашим направлениям, мудрых решений, терпения и отличного настроения каждый день."
]

WISHES_M = [
    "{name}, с днем рождения! Силы, уверенности, крутых результатов во всем и чтобы любые рабочие таски закрывались сходу. С праздником!",
    "С днем рождения, {name}! Желаем надежного тыла, крепких нервов, крутых целей и чтобы каждый проект приносил максимум профита."
]

WISHES_F = [
    "{name}, с днем рождения! Вдохновения, легкости, классных проектов и отличного настроения. Пусть все задуманное сбывается легко!",
    "С твоим днем, {name}! Желаем кайфовать от работы, не знать выгорания, гармонии во всем и пусть каждый день приносит только позитив."
]

WISHES_GENERAL = [
    "С днем рождения, {name}! Профессионального роста, кайфовой атмосферы в команде и побольше поводов для гордости за свои результаты.",
    "{name}, с праздником! Желаем не терять драйв, открывать новые горизонты и чтобы баланс между работой и жизнью всегда был идеальным."
]

def get_todays_birthdays():
    """Читает Excel через openpyxl и находит сегодняшних именинников."""
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        today = datetime.now()
        employees = []

        for row in ws.iter_rows(min_row=1, max_col=4):
            val_gender = row[0].value
            val_name = row[1].value
            val_position = row[2].value
            val_date = row[3].value

            if not val_name or not val_date:
                continue

            dob = None
            if isinstance(val_date, datetime):
                dob = val_date
            elif isinstance(val_date, str):
                try:
                    dob = datetime.strptime(val_date.strip(), "%d.%m.%Y")
                except ValueError:
                    pass

            if not dob:
                continue

            if dob.day == today.day and dob.month == today.month:
                is_formal = False

                cell_fill = row[1].fill
                if cell_fill and cell_fill.start_color and cell_fill.start_color.rgb == 'FFFFFF00':
                    is_formal = True

                # --- Логика обрезки ФИО ---
                full_name = str(val_name).strip()
                name_parts = full_name.split()

                if len(name_parts) >= 3:
                    wish_name = f"{name_parts[1]} {name_parts[2]}"  # Имя Отчество
                elif len(name_parts) == 2:
                    wish_name = name_parts[1]  # Только Имя
                else:
                    wish_name = full_name  # Если одно слово

                employees.append({
                    "full_name": full_name,
                    "wish_name": wish_name,
                    "position": str(val_position).strip() if val_position else "Сотрудник",
                    "gender": str(val_gender).strip().lower() if val_gender else None,
                    "is_formal": is_formal
                })

        return employees

    except Exception as e:
        logging.error(f"Ошибка при обработке Excel файла: {e}")
        return []

def generate_wish(name: str, gender: str, is_formal: bool) -> str:
    """Генерирует текст поздравления."""
    if is_formal:
        return random.choice(WISHES_FORMAL).format(name=name)

    if gender in ['м', 'муж', 'мужской']:
        pool = WISHES_M + WISHES_GENERAL
    elif gender in ['ж', 'жен', 'женский']:
        pool = WISHES_F + WISHES_GENERAL
    else:
        pool = WISHES_GENERAL

    return random.choice(pool).format(name=name)

async def send_birthday_reminders(bot: Bot):
    """Отправка уведомлений эйчару."""
    birthdays = get_todays_birthdays()

    if not birthdays:
        logging.info("Сегодня именинников нет.")
        return

    for emp in birthdays:
        # Передаем укороченное имя в генератор текста
        wish_text = generate_wish(emp['wish_name'], emp['gender'], emp['is_formal'])

        status_mark = "👑 **(Руководитель)**" if emp['is_formal'] else "👤"

        message = (
            f"🔔 **Напоминание:** Сегодня день рождения!\n"
            f"{status_mark} **{emp['full_name']}**\n"  
            f"💼 Должность: _{emp['position']}_\n\n"
            f"💡 **Текст для поздравления:**\n"
            f"{wish_text}"
        )

        try:
            await bot.send_message(chat_id=HR_CHAT_ID, text=message, parse_mode="Markdown")
            logging.info(f"Уведомление отправлено для: {emp['full_name']}")
        except Exception as e:
            logging.error(f"Не удалось отправить сообщение в чат {HR_CHAT_ID}: {e}")

async def main():
    logging.basicConfig(level=logging.INFO)

    bot = Bot(token=BOT_TOKEN)
    dp = Dispatcher()

    scheduler = AsyncIOScheduler(timezone="Asia/Almaty")

    scheduler.add_job(
        send_birthday_reminders,
        trigger="cron",
        hour=8,
        minute=0,
        kwargs={"bot": bot}
    )

    scheduler.start()
    logging.info("Бот запущен. Ожидание расписания...")

    try:
        await dp.start_polling(bot)
    finally:
        await bot.session.close()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logging.info("Бот остановлен.")